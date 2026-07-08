"""
Carga Microorganismo desde fuentes de referencia de patógenos humanos.

Fuentes combinadas (por defecto):
- Bartlett et al. 2022 — 1513 bacterias patógenas humanas (Microbiology Society)
- CZ ID Pathogen List 2024 — patógenos con taxon_id NCBI (bacterias, hongos, virus)
"""
from __future__ import annotations

import csv
import re
from io import StringIO
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction

from catalogos.management.catalog_sources import resolve_file
from laboratorio.models_microbiologia import Microorganismo

DEFAULT_BARTLETT_URL = (
    'https://github.com/padpadpadpad/bartlett_et_al_2022_human_pathogens/'
    'raw/master/data/bacteria_human_pathogens.xlsx'
)
DEFAULT_CZ_URL = (
    'https://raw.githubusercontent.com/icaromsc/CZ_ID_pathogen_list/main/'
    'CZ_ID_pathogen_list_2024.tsv'
)
BATCH_SIZE = 1000


def _normalize_name(name: str) -> str:
    return re.sub(r'\s+', ' ', name.strip().lower())


def _slug_codigo(genus: str, species: str, *, prefix: str = 'BAC') -> str:
    g = re.sub(r'[^a-zA-Z0-9]', '', genus)[:12]
    s = re.sub(r'[^a-zA-Z0-9]', '', species)[:20]
    code = f'{prefix}-{g}_{s}'.upper()
    return code[:40]


class Command(BaseCommand):
    help = 'Carga microorganismos desde Bartlett 2022 + CZ ID Pathogen List 2024'

    def add_arguments(self, parser):
        parser.add_argument('--bartlett-file', dest='bartlett_file', help='XLSX Bartlett local')
        parser.add_argument('--bartlett-url', default=DEFAULT_BARTLETT_URL)
        parser.add_argument('--cz-file', dest='cz_file', help='TSV CZ ID local')
        parser.add_argument('--cz-url', default=DEFAULT_CZ_URL)
        parser.add_argument(
            '--skip-bartlett',
            action='store_true',
            help='Solo cargar CZ ID (hongos/virus adicionales)',
        )
        parser.add_argument(
            '--skip-cz',
            action='store_true',
            help='Solo cargar Bartlett (bacterias)',
        )
        parser.add_argument('--clear', action='store_true')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit('Falta openpyxl. Instalá con: pip install openpyxl') from exc

        records: dict[str, dict] = {}

        if not options['skip_bartlett']:
            path = resolve_file(
                local_path=options.get('bartlett_file'),
                url=options['bartlett_url'],
                cache_name='bartlett_human_pathogens_2022.xlsx',
            )
            bartlett = self._parse_bartlett(openpyxl, path)
            self.stdout.write(f'Bartlett 2022: {len(bartlett)} bacterias')
            for rec in bartlett:
                records[_normalize_name(rec['nombre'])] = rec

        if not options['skip_cz']:
            cz_path = resolve_file(
                local_path=options.get('cz_file'),
                url=options['cz_url'],
                cache_name='cz_id_pathogen_list_2024.tsv',
            )
            cz_records = self._parse_cz(cz_path)
            self.stdout.write(f'CZ ID 2024: {len(cz_records)} patógenos')
            merged = 0
            added = 0
            for rec in cz_records:
                key = _normalize_name(rec['nombre'])
                if key in records:
                    existing = records[key]
                    existing['codigo'] = rec['codigo']
                    if rec.get('descripcion'):
                        existing['descripcion'] = rec['descripcion']
                    merged += 1
                else:
                    records[key] = rec
                    added += 1
            self.stdout.write(f'  Enriquecidos con NCBI: {merged}, nuevos (no Bartlett): {added}')

        final = list(records.values())
        self.stdout.write(f'Total únicos: {len(final)}')

        if options['dry_run']:
            for row in final[:5]:
                self.stdout.write(f"  {row['codigo']} — {row['nombre'][:60]}")
            return

        if options['clear']:
            deleted = Microorganismo.objects.count()
            Microorganismo.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Eliminados {deleted} microorganismos'))

        created = self._bulk_insert(final)
        total = Microorganismo.objects.count()
        self.stdout.write(
            self.style.SUCCESS(f'Carga completada: {created} insertados, {total} total')
        )

    def _parse_bartlett(self, openpyxl, path: str) -> list[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = 'Tab 6 Full List'
        if sheet not in wb.sheetnames:
            raise SystemExit(f'Hoja no encontrada: {sheet}')
        ws = wb[sheet]
        records: list[dict] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7:
                continue
            superkingdom = (row[0] or '').strip()
            phylum = (row[1] or '').strip()
            genus = (row[5] or '').strip()
            species = (row[6] or '').strip()
            if not genus or not species:
                continue
            nombre = f'{genus} {species}'
            codigo = _slug_codigo(genus, species)
            grupo = phylum or superkingdom or 'Bacteria'
            desc_parts = []
            if superkingdom:
                desc_parts.append(f'Superreino: {superkingdom}')
            if phylum:
                desc_parts.append(f'Filo: {phylum}')
            if len(row) > 8 and row[8]:
                desc_parts.append(f'Estado: {row[8]}')
            desc_parts.append('Fuente: Bartlett et al. 2022 (Microbiology 168:001269)')

            records.append(
                {
                    'codigo': codigo,
                    'nombre': nombre[:200],
                    'genero': genus[:120],
                    'especie': species[:120],
                    'grupo': grupo[:80],
                    'descripcion': '. '.join(desc_parts),
                    'activo': True,
                }
            )
        return records

    def _parse_cz(self, path: str) -> list[dict]:
        content = Path(path).read_text(encoding='utf-8-sig')
        reader = csv.DictReader(content.splitlines(), delimiter='\t')
        records: list[dict] = []

        for row in reader:
            nombre = (row.get('taxon_name') or '').strip()
            taxon_id = (row.get('taxon_id') or '').strip()
            if not nombre or not taxon_id:
                continue
            parts = nombre.split(None, 1)
            genero = parts[0] if parts else ''
            especie = parts[1] if len(parts) > 1 else nombre

            records.append(
                {
                    'codigo': taxon_id[:40],
                    'nombre': nombre[:200],
                    'genero': genero[:120],
                    'especie': especie[:120],
                    'grupo': self._infer_grupo(nombre),
                    'descripcion': f'NCBI Taxonomy ID: {taxon_id}. Fuente: CZ ID Pathogen List 2024.',
                    'activo': True,
                }
            )
        return records

    @staticmethod
    def _infer_grupo(nombre: str) -> str:
        lower = nombre.lower()
        virus_markers = ('virus', 'viridae', 'phage', 'herpes', 'influenza', 'coronavirus')
        fungus_markers = (
            'candida',
            'aspergillus',
            'cryptococcus',
            'pneumocystis',
            'histoplasma',
            'coccidioides',
            ' mucor',
            ' fusarium',
        )
        if any(m in lower for m in virus_markers):
            return 'Virus'
        if any(m in lower for m in fungus_markers):
            return 'Hongos'
        return 'Bacteria'

    def _bulk_insert(self, records: list[dict]) -> int:
        existing = set(Microorganismo.objects.values_list('codigo', flat=True))
        to_create = [
            Microorganismo(**row) for row in records if row['codigo'] not in existing
        ]
        created = 0
        with transaction.atomic():
            for i in range(0, len(to_create), BATCH_SIZE):
                batch = to_create[i : i + BATCH_SIZE]
                Microorganismo.objects.bulk_create(batch, ignore_conflicts=True)
                created += len(batch)
                if (i + BATCH_SIZE) % 2000 == 0 or i + BATCH_SIZE >= len(to_create):
                    self.stdout.write(f'  Insertados {min(i + BATCH_SIZE, len(to_create))}…')
        return created
