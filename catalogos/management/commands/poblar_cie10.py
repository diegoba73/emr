"""
Carga la tabla DiagnosticoCIE10 desde CIE-10-ES 5ª edición (2024).

Fuente oficial: Ministerio de Sanidad de España — Tabla de referencia de diagnósticos.
"""
from __future__ import annotations

import os
import tempfile
from pathlib import Path

import requests
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from catalogos.models import DiagnosticoCIE10

DEFAULT_XLSX_URL = (
    'https://www.sanidad.gob.es/estadEstudios/estadisticas/normalizacion/CIE10/2024/'
    'Diagnosticos_ES2024_TablaReferencia_30_06_2023_9096243130459179657.xlsx'
)
DEFAULT_SHEET = 'ES2024 Completa + Marcadores'
BATCH_SIZE = 2000


class Command(BaseCommand):
    help = (
        'Descarga y carga CIE-10-ES 2024 (Tabla de referencia diagnósticos) '
        'en catalogos.DiagnosticoCIE10'
    )

    def add_arguments(self, parser):
        parser.add_argument(
            '--url',
            default=DEFAULT_XLSX_URL,
            help='URL del XLSX oficial CIE-10-ES 2024',
        )
        parser.add_argument(
            '--file',
            dest='xlsx_file',
            help='Ruta local al XLSX (omite descarga)',
        )
        parser.add_argument(
            '--sheet',
            default=DEFAULT_SHEET,
            help='Hoja a importar (por defecto: completa con jerarquía)',
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Eliminar diagnósticos existentes antes de cargar',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Solo parsear y mostrar estadísticas, sin escribir en BD',
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit(
                'Falta openpyxl. Instalá con: pip install openpyxl'
            ) from exc

        xlsx_path = options.get('xlsx_file')
        if xlsx_path:
            if not os.path.exists(xlsx_path):
                self.stderr.write(self.style.ERROR(f'No existe: {xlsx_path}'))
                return
        else:
            xlsx_path = self._download_xlsx(options['url'])

        records = self._parse_xlsx(openpyxl, xlsx_path, options['sheet'])
        self.stdout.write(f'Registros parseados (códigos finales): {len(records)}')

        if options['dry_run']:
            for sample in records[:5]:
                self.stdout.write(f"  {sample['codigo']} — {sample['descripcion'][:60]}")
            return

        if options['clear']:
            deleted = DiagnosticoCIE10.objects.count()
            DiagnosticoCIE10.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Eliminados {deleted} diagnósticos previos'))

        created = self._bulk_insert(records)
        total = DiagnosticoCIE10.objects.count()
        self.stdout.write(
            self.style.SUCCESS(
                f'Carga CIE-10-ES 2024 completada: {created} insertados, '
                f'{total} total en base de datos'
            )
        )

    def _download_xlsx(self, url: str) -> str:
        self.stdout.write(f'Descargando CIE-10-ES desde:\n  {url}')
        cache_dir = Path(settings.BASE_DIR) / 'data'
        cache_dir.mkdir(exist_ok=True)
        cache_path = cache_dir / 'cie10_es2024_diagnosticos.xlsx'

        try:
            response = requests.get(url, timeout=120)
            response.raise_for_status()
            cache_path.write_bytes(response.content)
            self.stdout.write(self.style.SUCCESS(f'Guardado en {cache_path}'))
            return str(cache_path)
        except Exception as exc:
            self.stderr.write(self.style.WARNING(f'Descarga fallida ({exc}); intento temporal'))
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            response = requests.get(url, timeout=120)
            response.raise_for_status()
            tmp.write(response.content)
            tmp.close()
            return tmp.name

    def _parse_xlsx(self, openpyxl, path: str, sheet_name: str) -> list[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SystemExit(
                f'Hoja "{sheet_name}" no encontrada. Disponibles: {", ".join(wb.sheetnames)}'
            )
        ws = wb[sheet_name]

        current_capitulo = ''
        current_bloque = ''
        current_enfermedad = ''
        records: list[dict] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = (row[0] or '').strip()
            descripcion = (row[1] or '').strip()
            nodo_final = row[2] if len(row) > 2 else None

            if not codigo or not descripcion:
                continue

            if codigo.startswith('Cap.'):
                current_capitulo = descripcion
                continue

            if '-' in codigo and '.' not in codigo and len(codigo) <= 7:
                current_bloque = descripcion
                continue

            if nodo_final == 0 and '.' not in codigo and len(codigo) <= 3:
                current_enfermedad = descripcion
                continue

            if nodo_final != 1:
                continue

            categoria = codigo.split('.')[0] if '.' in codigo else codigo[:3]
            records.append(
                {
                    'codigo': codigo[:10],
                    'descripcion': descripcion[:5000],
                    'categoria': (current_bloque or categoria)[:100],
                    'capitulo': current_capitulo[:100] if current_capitulo else None,
                    'enfermedad': (current_enfermedad or current_bloque or categoria)[:200],
                    'tipo_enfermedad': categoria[:200],
                    'activo': True,
                }
            )

        wb.close()
        return records

    def _bulk_insert(self, records: list[dict]) -> int:
        existing = set(DiagnosticoCIE10.objects.values_list('codigo', flat=True))
        to_create = [
            DiagnosticoCIE10(**row)
            for row in records
            if row['codigo'] not in existing
        ]
        created = 0
        with transaction.atomic():
            for i in range(0, len(to_create), BATCH_SIZE):
                batch = to_create[i : i + BATCH_SIZE]
                DiagnosticoCIE10.objects.bulk_create(batch, ignore_conflicts=True)
                created += len(batch)
                if (i + BATCH_SIZE) % 10000 == 0 or i + BATCH_SIZE >= len(to_create):
                    self.stdout.write(f'  Insertados {min(i + BATCH_SIZE, len(to_create))}…')

        # Actualizar existentes (por si no se usó --clear)
        if existing:
            by_code = {r['codigo']: r for r in records}
            updates = []
            for obj in DiagnosticoCIE10.objects.filter(codigo__in=existing):
                data = by_code.get(obj.codigo)
                if not data:
                    continue
                obj.descripcion = data['descripcion']
                obj.categoria = data['categoria']
                obj.capitulo = data['capitulo']
                obj.enfermedad = data['enfermedad']
                obj.tipo_enfermedad = data['tipo_enfermedad']
                obj.activo = True
                updates.append(obj)
            if updates:
                DiagnosticoCIE10.objects.bulk_update(
                    updates,
                    ['descripcion', 'categoria', 'capitulo', 'enfermedad', 'tipo_enfermedad', 'activo'],
                    batch_size=BATCH_SIZE,
                )
        return created
