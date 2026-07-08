"""
Carga ProcedimientoCatalogo desde CIE-10-ES Procedimientos 5ª ed. 2024 (ICD-10-PCS-ES).

Fuente: Ministerio de Sanidad de España — Tabla de referencia de procedimientos.
"""
from __future__ import annotations

from django.core.management.base import BaseCommand
from django.db import transaction

from catalogos.management.catalog_sources import resolve_file
from catalogos.models import ProcedimientoCatalogo

DEFAULT_XLSX_URL = (
    'https://www.sanidad.gob.es/estadEstudios/estadisticas/normalizacion/CIE10/2024/'
    'Procedimientos_ES2024_TablaReferencia_30062023_5537663830978566667.xlsx'
)
DEFAULT_SHEET = 'ES2024 Completa + Marcadores'
BATCH_SIZE = 2000


class Command(BaseCommand):
    help = 'Descarga y carga CIE-10-ES 2024 procedimientos en ProcedimientoCatalogo'

    def add_arguments(self, parser):
        parser.add_argument('--url', default=DEFAULT_XLSX_URL)
        parser.add_argument('--file', dest='xlsx_file')
        parser.add_argument('--sheet', default=DEFAULT_SHEET)
        parser.add_argument('--clear', action='store_true')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:
            raise SystemExit('Falta openpyxl. Instalá con: pip install openpyxl') from exc

        path = resolve_file(
            local_path=options.get('xlsx_file'),
            url=options['url'],
            cache_name='cie10_es2024_procedimientos.xlsx',
        )
        records = self._parse(openpyxl, path, options['sheet'])
        self.stdout.write(f'Registros parseados: {len(records)}')

        if options['dry_run']:
            for row in records[:5]:
                self.stdout.write(f"  {row['nombre'][:80]}")
            return

        if options['clear']:
            deleted = ProcedimientoCatalogo.objects.count()
            ProcedimientoCatalogo.objects.all().delete()
            self.stdout.write(self.style.WARNING(f'Eliminados {deleted} procedimientos'))

        created = self._bulk_insert(records)
        total = ProcedimientoCatalogo.objects.count()
        self.stdout.write(
            self.style.SUCCESS(f'Carga completada: {created} insertados, {total} total')
        )

    def _parse(self, openpyxl, path: str, sheet_name: str) -> list[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f'Hoja no encontrada: {sheet_name}')
        ws = wb[sheet_name]
        records: list[dict] = []
        seen_names: set[str] = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = (row[0] or '').strip()
            descripcion = (row[1] or '').strip()
            if not codigo or not descripcion:
                continue
            nombre = descripcion[:255]
            if nombre in seen_names:
                nombre = f'{nombre[:240]} ({codigo})'[:255]
            seen_names.add(nombre)
            records.append(
                {
                    'nombre': nombre,
                    'descripcion': f'{codigo} — {descripcion}'[:5000],
                    'activo': True,
                }
            )
        wb.close()
        return records

    def _bulk_insert(self, records: list[dict]) -> int:
        existing = set(ProcedimientoCatalogo.objects.values_list('nombre', flat=True))
        to_create = [
            ProcedimientoCatalogo(**row)
            for row in records
            if row['nombre'] not in existing
        ]
        created = 0
        with transaction.atomic():
            for i in range(0, len(to_create), BATCH_SIZE):
                batch = to_create[i : i + BATCH_SIZE]
                ProcedimientoCatalogo.objects.bulk_create(batch, ignore_conflicts=True)
                created += len(batch)
                if (i + BATCH_SIZE) % 20000 == 0 or i + BATCH_SIZE >= len(to_create):
                    self.stdout.write(f'  Insertados {min(i + BATCH_SIZE, len(to_create))}…')
        return created
